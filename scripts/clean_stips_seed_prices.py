"""
clean_stips_seed_prices.py
--------------------------
Convert STIPS seed-price XLS files to a normalized long-format CSV.

Input:  data/raw/downloaded_xls/*.xls
Output: data/processed/stips_seed_prices_clean.csv
        (Year, Crop, Packaging, City, Seed price (RSD))
"""

import os
import re
import unicodedata

import openpyxl
import pandas as pd
import xlrd

from paths import DOWNLOADED_XLS_DIR, STIPS_SEED_PRICES_CLEAN

CROP_CATEGORY_MAP = {
    "KROMPIR": None,
    "krompir": None,
    "KUKURUZ": "Corn",
    "kukuruz": "Corn",
    "KURURUZ": "Corn",
    "PSENICA": "Wheat",
    "PŠENICA": "Wheat",
    "pšenica": "Wheat",
    "JECAM": "Barley",
    "JEČAM": "Barley",
    "ječam": "Barley",
}

KNOWN_CITIES = {
    "Beograd",
    "Čačak",
    "Kragujevac",
    "Kraljevo",
    "Loznica",
    "Niš",
    "Pirot",
    "Požarevac",
    "Smederevo",
    "Vranje",
    "Zaječar",
    "Kikinda",
    "Novi Sad",
    "Pančevo",
    "Sombor",
    "Sremska Mitrovica",
    "Subotica",
    "Zrenjanin",
}

def _read_xls(file_path):
    workbook = xlrd.open_workbook(file_path)
    all_sheets = []
    for index in range(workbook.nsheets):
        worksheet = workbook.sheet_by_index(index)
        rows = [worksheet.row_values(row_idx) for row_idx in range(worksheet.nrows)]
        all_sheets.append((worksheet.name, rows))
    return all_sheets


def _read_xlsx(file_path):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    all_sheets = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        all_sheets.append((sheet_name, rows))
    return all_sheets


def read_all_sheets(file_path):
    if file_path.lower().endswith(".xlsx"):
        return _read_xlsx(file_path)
    return _read_xls(file_path)


def _clean_cell(value):
    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKC", text).strip()
    return text


def find_city_header(rows):
    for row_idx, row in enumerate(rows):
        cleaned = [_clean_cell(value) for value in row]
        city_columns = {
            col_idx: cleaned[col_idx]
            for col_idx in range(len(cleaned))
            if cleaned[col_idx] in KNOWN_CITIES
        }
        if len(city_columns) >= 5:
            return row_idx, city_columns
    return None, {}


def parse_filename(filename):
    """Parse semenski_april_2010.xls -> ('april', 2010)."""
    base = os.path.splitext(filename)[0]
    base = re.sub(r"_\d{1,2}$", "", base)
    parts = base.split("_")
    if len(parts) >= 3:
        month = parts[1]
        try:
            year = int(parts[2])
        except ValueError:
            year = None
        return month, year
    return None, None


def parse_sheet(sheet_name, rows, year, month):
    header_idx, city_columns = find_city_header(rows)
    if header_idx is None:
        return None

    records = []
    for row in rows[header_idx + 1 :]:
        if not row:
            continue

        product = _clean_cell(row[0] if len(row) > 0 else "")
        producer = _clean_cell(row[1] if len(row) > 1 else "")
        packaging = _clean_cell(row[2] if len(row) > 2 else "")

        if not product:
            continue

        product = re.sub(r"\s+", " ", product)

        has_price = False
        for col_idx in city_columns:
            if col_idx < len(row):
                value = row[col_idx]
                if value is not None and value != "":
                    try:
                        float(value)
                        has_price = True
                        break
                    except (ValueError, TypeError):
                        pass

        if not has_price:
            continue

        for col_idx, city in city_columns.items():
            price = None
            if col_idx < len(row):
                value = row[col_idx]
                if value is not None and value != "":
                    try:
                        price = float(value)
                    except (ValueError, TypeError):
                        pass

            records.append(
                {
                    "Year": year,
                    "Month": month,
                    "Category": sheet_name.strip(),
                    "Product": product,
                    "Producer": producer,
                    "Packaging": packaging,
                    "City": city,
                    "Seed price (RSD)": price,
                }
            )

    if not records:
        return None
    return pd.DataFrame(records)


def postprocess(data):
    """
    - normalize Category to English crop names
    - drop potato and unknown categories
    - drop Month, Product and Producer columns
    - drop rows with missing seed prices
    """
    data = data.copy()
    data["Category"] = data["Category"].str.strip()

    unknown_categories = set(data["Category"].unique()) - set(CROP_CATEGORY_MAP.keys())
    if unknown_categories:
        print(f"  WARNING: unknown Category values dropped: {unknown_categories}")

    data["Crop"] = data["Category"].map(CROP_CATEGORY_MAP)
    data = data.dropna(subset=["Crop"])

    data = data.drop(columns=["Category", "Month", "Product", "Producer"])
    data = data.dropna(subset=["Seed price (RSD)"])
    return data


def main():
    file_list = sorted(os.listdir(DOWNLOADED_XLS_DIR))
    all_dataframes = []

    for filename in file_list:
        if not (filename.lower().endswith(".xls") or filename.lower().endswith(".xlsx")):
            continue

        month, year = parse_filename(filename)
        if year is None:
            continue

        file_path = os.path.join(DOWNLOADED_XLS_DIR, filename)
        try:
            sheets = read_all_sheets(file_path)
        except Exception as exc:
            print(f"  ERROR reading {filename}: {exc}")
            continue

        file_dataframes = []
        for sheet_name, rows in sheets:
            dataframe = parse_sheet(sheet_name, rows, year, month)
            if dataframe is not None and not dataframe.empty:
                file_dataframes.append(dataframe)

        total_rows = sum(len(frame) for frame in file_dataframes)
        if file_dataframes:
            all_dataframes.append(pd.concat(file_dataframes, ignore_index=True))
            print(f"  OK  {filename:50s}  {total_rows:>6} rows  ({len(file_dataframes)} sheet(s))")
        else:
            print(f"  WARN {filename}: no data extracted")

    if not all_dataframes:
        print("\nNo data to process. Check data/raw/downloaded_xls/.")
        return

    result = pd.concat(all_dataframes, ignore_index=True)
    result = postprocess(result)
    result = result.sort_values(["Year", "Crop", "Packaging", "City"]).reset_index(drop=True)

    result.to_csv(STIPS_SEED_PRICES_CLEAN, index=False, encoding="utf-8-sig")
    print(
        f"\nSaved  {STIPS_SEED_PRICES_CLEAN}  ({len(result):,} rows, "
        f"{result['Crop'].nunique()} crops, "
        f"{result['Year'].nunique()} years)"
    )


if __name__ == "__main__":
    main()
